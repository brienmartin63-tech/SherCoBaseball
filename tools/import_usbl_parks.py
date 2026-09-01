#!/usr/bin/env python3
"""Convert Brien's 28x28 USBL park workbook into app-ready JSON.

Excel stores the visual grid from upper-left to lower-right. SherCo coordinates
run from home plate outward, so worksheet cell A1 is coordinate 28-28 and cell
AB28 is coordinate 1-1.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


COLOR_TO_TERRAIN = {
    "99CC00": "field",
    "FFFF00": "beyondFence",
    "FFCC00": "dirt",
    "0066CC": "field",
}


def rgb(cell) -> str:
    value = (cell.fill.fgColor.rgb or "").upper()
    return value[-6:]


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def import_workbook(source: Path) -> list[dict]:
    workbook = load_workbook(source, data_only=True)
    parks = []
    for sheet in workbook.worksheets:
        cells = []
        fielders = []
        for excel_row in range(1, 29):
            row = []
            for excel_col in range(1, 29):
                cell = sheet.cell(excel_row, excel_col)
                color = rgb(cell)
                terrain = COLOR_TO_TERRAIN.get(color, "special")
                row.append(terrain)
                if color == "0066CC" and cell.value:
                    fielders.append({
                        "position": str(cell.value),
                        "at": {"row": 29 - excel_row, "column": 29 - excel_col},
                    })
            cells.append(row)

        name = str(sheet["A30"].value).strip()
        home_line = str(sheet["A31"].value).strip()
        team = home_line.replace("Home of the ", "", 1)
        location = str(sheet["A32"].value).strip().title()
        parks.append({
            "id": slug(name),
            "name": name.title(),
            "team": team,
            "location": location,
            "dimensions": 28,
            "cells": cells,
            "fielders": sorted(fielders, key=lambda f: f["position"]),
            "sourceSheet": sheet.title,
        })
    return parks


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("usage: import_usbl_parks.py SOURCE.xlsx DESTINATION.json")
    source, destination = map(Path, sys.argv[1:])
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(json.dumps(import_workbook(source), indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
